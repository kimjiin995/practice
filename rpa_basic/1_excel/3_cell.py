from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["b1"] = 4
ws["b2"] = 5
ws["b3"] = 6
ws["c1"] = 7
ws["c2"] = 1
ws["c3"] = 1

ws.cell(column=1, row=1, value=1)
ws.cell(column=1, row=2, value=1)
ws.cell(column=1, row=3, value=1)
ws.cell(column=2, row=1, value=1)
ws.cell(column=2, row=2, value=1)

# 출력할때
print(ws["A1"])
print(ws["a1"].value)
print(ws.cell(column=1, row=1))
print(ws.cell(column=1, row=1).value)
print(ws["A10"].value) # 값이 없을 땐 None

# 반복문
index = 1
for y in range(1, 11):
    for x in range(1, 11):
        ws.cell(row=y, column=x, value=index)
        index += 1

wb.save("example.xlsx")
wb.close()