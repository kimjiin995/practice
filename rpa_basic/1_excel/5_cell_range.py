from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active
ws.append(["번호", "영어", "수학"])

for i in range(1, 10):
    ws.append([str(i+1)+"번", randint(0, 100), randint(0,100)])

col_B = ws["B"] # 영어 column만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_ranges =ws["B":"C"]
# for cols in col_ranges:
#     for cell in cols:
#         print(cell.value)

# col_ranges =ws["B":"C"]
# for cols in col_ranges:
#     for cell in cols:
#         cell.value=1

# row_range = ws[2:5]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         print(cell.coordinate, end=" ")
#     print()

# row_range = ws[2:ws.max_row]
# for row in row_range:
#     for cell in row:
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end=" ")
#     print()

# row_range = ws[2:ws.max_row]
# for row in row_range:
#     for cell in row:
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# print(tuple(ws.rows))
# print(tuple(ws.columns))
# print(tuple(ws.iter_rows()))
# print(tuple(ws.iter_cols()))

# for row in ws.rows:
#     print(row)

# for row in ws.rows:
#     print(row[1].value) # row당 첫번째 컬럼 가져옴

# for col in ws.columns:
#     print(col)

# for col in ws.columns:
#     print(col[1].value, end=" ") # column당 첫번째 row 가져옴

# for row in ws.rows:
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

# for col in ws.columns:
#     for cell in col:
#         print(cell.value, end=" ")
#     print()

# for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5):
#     print(row[0].value, row[1].value, row[2].value)

# for col in ws.iter_cols(min_row=2, max_row=4, min_col=2, max_col=5):
#     print(col[0].value, col[1].value, col[2].value)

# for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5):
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

# for col in ws.iter_cols(min_row=2, max_row=4, min_col=2, max_col=5):
#     for cell in col:
#         print(cell.value, end=" ")
#     print()

wb.save("sample2.xlsx")