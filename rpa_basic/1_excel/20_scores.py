from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# excel에 데이터 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores:
    ws.append(s)

# 퀴즈2 점수를 10으로 수정
for cell in ws["D"]:
    if cell.row == 1:
        continue
    cell.value = 10

# H열에 총점(SUM), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
ws["H1"] = "총점"
ws["I1"] = "성적"

# for 문 여러개 돌릴 수 있으니까 ws.cell()형식으로 하기
for idx, score in enumerate(scores, 2):

    # H열에 총점 함수식 넣기
    ws.cell(row=idx, column=8, value=f"=SUM(B{idx}:G{idx})")
    # evaluate 안되니까 sum값 따로 구하기
    sum_val = sum(score[1:]) - score[3] + 10

    # I열 성적 정보 구하기
    if sum_val >= 90:
        grade = "A" 
    elif sum_val >= 80:
        grade = "B" 
    elif sum_val >= 70:
        grade = "C" 
    else:
        grade ="D" 
    
    # 출석이 5미만인 학생은 총점 상관없이 F
    if score[1] < 5:
        grade = "F"
    
    # I열에 성적 값 넣기
    ws.cell(row=idx, column=9, value=grade)

    wb.save("scores.xlsx")


