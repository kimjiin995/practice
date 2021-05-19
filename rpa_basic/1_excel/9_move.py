from openpyxl import load_workbook

wb = load_workbook("sample2.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"
# wb.save("sample2_move.xlsx")

ws.move_range("C1:C11", rows=4, cols=-1)
wb.save("sample2_move2.xlsx")