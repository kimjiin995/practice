from openpyxl import load_workbook

wb = load_workbook("quiz.xlsx", data_only=True)
ws = wb.active

ws["I1"] = "성적"
for cells in ws["H2":"I11"]:
    if cells[0].value >= 90:
        cells[1].value = "A"
    elif cells[0].value >= 80:
        cells[1].value = "B"
    elif cells[0].value >= 70:
        cells[1].value = "C"
    else:
        cells[1].value = "D"

# 출석이 5미만인 학생은 총점 상관없이 F
for cell in ws["B"]:
    if cell.row == 1:
        continue
    if int(cell.value) < 5:
        ws[f"I{cell.row}"].value = "F"
        

wb.save("score.xlsx")

