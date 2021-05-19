from openpyxl import load_workbook
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
wb = load_workbook("sample2.xlsx")
ws = wb.active

ws.column_dimensions["A"].width = 20
ws.row_dimensions[1].height = 50

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 폰트
a1.font = Font(color="FF0000", size=20, name="arial", italic=True, bold=True, strike=True, underline="single")

# 테두리적용
thin_border = Border(left=Side(style="thin", color="FF0000"), right=Side(style="thin", color="FF0000"), top=Side(style="thin", color="FF0000"), bottom=Side(style="thin", color="FF0000"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border


green_fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경색 채우기

for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") # 정렬
        if cell in ws["A"]:
            continue
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = green_fill # 배경색 채우기
            cell.font = Font(color="FF0000")
    
ws.freeze_panes = "B2" # 틀 고정

wb.save("sample2_style.xlsx")