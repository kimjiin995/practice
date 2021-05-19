from openpyxl import load_workbook

wb = load_workbook("sample2.xlsx")
ws = wb.active

# 높은 점수 찾기
for row in ws.iter_rows(min_row=2):
    if row[1].value > 60:
        print(f"{row[0].value}번은 {row[1].value}점수를 기록해 영어 경시대회 후보입니다.")

# for row in ws[2:ws.max_row]:
#     if row[1].value > 60:
#         print(f"{row[0].value}번은 {row[1].value}점수를 기록해 영어 경시대회 후보입니다.")

# 영어를 컴퓨터로 바꾸기
for cell in ws[1]:
    if cell.value == "영어":
        cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")



