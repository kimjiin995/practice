from openpyxl import Workbook

wb = Workbook()
# ws = wb.active
ws = wb.create_sheet()
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff6666"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2) # 인덱스에 시트 생성
new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근가능 = sheet의 title을 알면 접근가능하단 말

print(wb.sheetnames)

new_ws["A1"] = "Test" 
target = wb.copy_worksheet(new_ws) # 시트 복사 (뒤에 붙음)
target.title = "Copied Sheet"

wb.save("example.xlsx")
wb.close()