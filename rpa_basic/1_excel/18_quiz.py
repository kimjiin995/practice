from openpyxl import load_workbook
from xlcalculator import ModelCompiler
from xlcalculator import Model
from xlcalculator import Evaluator
# import pandas as pd

wb = load_workbook("quiz_data.xlsx", data_only=True)
ws = wb.active

# 퀴즈2 점수를 10으로 수정
for cell in ws["D"]:
    if cell.row == 1:
        continue
    cell.value = 10

# H열에 총점(SUM), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
ws["H1"] = "총점"
i = 2
for cell in ws["H"]:
    
    if cell.row == 1:
        continue
    
    cell.value = f"=SUM(B{i}:G{i})"
    i += 1

wb.save("quiz.xlsx")
# pd.read_excel("quiz.xlsx")

# filename = "quiz.xlsx"
# compiler = ModelCompiler()
# new_model = compiler.read_and_parse_archive(filename)
# evaluator = Evaluator(new_model)
# for i in range(2, 13):
#     evaluator.evaluate(f'현재까지 작성된 최종 성적 데이터!H{i}')

# wb = load_workbook("quiz.xlsx", data_only=True)
# wb.save("quiz.xlsx")
# wb = load_workbook("quiz.xlsx", data_only=True)
# ws = wb.active

# ws["I1"] = "성적"
# for cells in ws["H2":"I11"]:
#     if cells[0].value >= 90:
#         cell[1].value = "A"
#     elif cells[0].value >= 80:
#         cell[1].value = "B"
#     elif cells[0].value >= 70:
#         cell[1].value = "C"
#     else:
#         cell[1].value = "D"


