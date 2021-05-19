from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

print(len(data))
for s in scores:
    ws.append(s)

wb.save("scores.xlsx")

#  from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active

# ws.title = "현재까지 작성된 최종 성적 데이터"
# data = ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트",
# 1,10,8,5,14,26,12,
# 2,7,3,7,15,24,18,
# 3,9,5,8,8,12,4,
# 4,7,8,7,17,21,18,
# 5,7,8,7,16,25,15,
# 6,3,5,8,8,17,0,
# 7,4,9,10,16,27,18,
# 8,6,6,6,15,19,17,
# 9,10,10,9,19,30,19,
# 10,9,8,8,20,25,20]

# print(len(data))
# index = 0
# for row in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col=7):
#     for cell in row:
#         cell.value = data[index]
#         index += 1

# wb.save("quiz_data.xlsx")